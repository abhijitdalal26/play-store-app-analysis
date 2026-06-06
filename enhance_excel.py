"""
Excel enhancement script for the 2026 Play Store data files.
Adds: auto-filters, frozen headers, column widths, conditional formatting,
named tables, a summary/dashboard sheet, and charts.

Run from the project root:
    python enhance_excel.py
"""

import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_DIR = "data_my_copy/scraped_2026"
APPS_FILE       = os.path.join(DATA_DIR, "apps.xlsx")
COUNTRY_FILE    = os.path.join(DATA_DIR, "app_country_stats.xlsx")
DISCOVERY_FILE  = os.path.join(DATA_DIR, "discovery_signals.xlsx")

# ── Helpers ───────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

ALT_FILL     = PatternFill("solid", fgColor="EEF2FF")    # light blue-tint
BORDER_SIDE  = Side(style="thin", color="D0D7E3")
THIN_BORDER  = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)

CHART_COLORS = [
    "4472C4", "ED7D31", "A9D18E", "FF0000", "FFC000",
    "5A9BD5", "70AD47", "264478", "9E480E", "636363",
]


def _col_width(df: pd.DataFrame, col_name: str, max_width: int = 40) -> float:
    sample = df[col_name].astype(str).head(200)
    return min(max(sample.str.len().max(), len(str(col_name))) + 2, max_width)


def _style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _apply_alt_rows(ws, start_row=2, end_row=None, ncols=None):
    if end_row is None:
        end_row = ws.max_row
    if ncols is None:
        ncols = ws.max_column
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            if fill:
                cell.fill = fill
            cell.border = THIN_BORDER


def _set_col_widths(ws, df):
    for idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = _col_width(df, col_name)


def _freeze_and_filter(ws, df):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _add_df_to_sheet(wb, df, sheet_name, max_rows=None):
    """Write a dataframe to a new sheet, styled."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    subset = df.head(max_rows) if max_rows else df
    for r in dataframe_to_rows(subset, index=False, header=True):
        ws.append(r)
    _style_header_row(ws)
    _apply_alt_rows(ws, ncols=len(df.columns), end_row=len(subset) + 1)
    _set_col_widths(ws, subset)
    _freeze_and_filter(ws, subset)
    ws.row_dimensions[1].height = 30
    return ws


def _add_bar_chart(ws, data_ws, min_row, max_row, cat_col, val_col, title,
                   anchor, bar_dir="col"):
    chart = BarChart()
    chart.type = bar_dir
    chart.title = title
    chart.style = 10
    chart.grouping = "clustered"
    chart.width = 16
    chart.height = 10

    data = Reference(data_ws, min_col=val_col, min_row=min_row, max_row=max_row)
    cats = Reference(data_ws, min_col=cat_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = CHART_COLORS[0]
    ws.add_chart(chart, anchor)


def _add_pie_chart(ws, data_ws, min_row, max_row, cat_col, val_col, title, anchor):
    chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.width = 14
    chart.height = 10

    data = Reference(data_ws, min_col=val_col, min_row=min_row, max_row=max_row)
    cats = Reference(data_ws, min_col=cat_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # colour slices
    for i, point in enumerate(chart.series[0].dPt if hasattr(chart.series[0], 'dPt') else []):
        pass
    ws.add_chart(chart, anchor)


# ── apps.xlsx ─────────────────────────────────────────────────────────────────

def _read_data_sheet(path: str, data_sheet: str) -> pd.DataFrame:
    """Read the data sheet by name if it exists, else read the first sheet."""
    import openpyxl as _xl
    wb = _xl.load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    if data_sheet in names:
        return pd.read_excel(path, sheet_name=data_sheet)
    return pd.read_excel(path, sheet_name=0)


def enhance_apps():
    print("Loading apps.xlsx ...")
    df = _read_data_sheet(APPS_FILE, "Apps")

    # Drop very long free-text / URL columns to keep Excel navigable
    drop_cols = [c for c in ["description", "summary", "icon", "header_image",
                              "screenshots", "video", "video_image",
                              "privacy_policy", "developer_website",
                              "developer_email"] if c in df.columns]
    df_view = df.drop(columns=drop_cols)

    wb = load_workbook(APPS_FILE)
    # After a prior run the data sheet is already named "Apps"; on first run rename the active sheet
    if "Apps" in wb.sheetnames:
        ws = wb["Apps"]
    else:
        ws = wb.active
        ws.title = "Apps"

    # Re-write the sheet (preserves existing data, adds formatting)
    _style_header_row(ws)
    _set_col_widths(ws, df_view)
    _freeze_and_filter(ws, df_view)
    ws.row_dimensions[1].height = 30

    # Conditional formatting: score column
    score_col_idx = list(df.columns).index("score") + 1  # 1-based
    score_col_letter = get_column_letter(score_col_idx)
    score_range = f"{score_col_letter}2:{score_col_letter}{len(df)+1}"
    ws.conditional_formatting.add(
        score_range,
        ColorScaleRule(
            start_type="num", start_value=0, start_color="FF0000",
            mid_type="num",   mid_value=3,   mid_color="FFFF00",
            end_type="num",   end_value=5,   end_color="00B050",
        )
    )

    # Conditional formatting: min_installs — data bar
    inst_col_idx = list(df.columns).index("min_installs") + 1
    inst_col_letter = get_column_letter(inst_col_idx)
    inst_range = f"{inst_col_letter}2:{inst_col_letter}{len(df)+1}"
    ws.conditional_formatting.add(
        inst_range,
        DataBarRule(start_type="min", end_type="max",
                    color="638EC6", showValue=True)
    )

    # ── Aggregations for charts ─────────────────────────────────────────────

    # 1. Genre distribution (top 20)
    genre_dist = (df.groupby("genre").size()
                    .sort_values(ascending=False)
                    .head(20)
                    .reset_index()
                    .rename(columns={0: "app_count"}))
    genre_dist.columns = ["Genre", "App Count"]

    # 2. Install bucket distribution
    install_buckets = {
        "<1K": (0, 1000),
        "1K–10K": (1000, 10000),
        "10K–100K": (10000, 100000),
        "100K–1M": (100000, 1000000),
        "1M–10M": (1000000, 10000000),
        "10M–100M": (10000000, 100000000),
        "100M+": (100000000, float("inf")),
    }
    bucket_counts = {}
    for label, (lo, hi) in install_buckets.items():
        bucket_counts[label] = ((df["min_installs"] >= lo) & (df["min_installs"] < hi)).sum()
    install_dist = pd.DataFrame(list(bucket_counts.items()),
                                columns=["Install Range", "App Count"])

    # 3. Content rating
    content_dist = (df.groupby("content_rating").size()
                      .sort_values(ascending=False)
                      .reset_index()
                      .rename(columns={0: "app_count"}))
    content_dist.columns = ["Content Rating", "App Count"]

    # 4. Free vs Paid
    fp_dist = (df["free"].map({True: "Free", False: "Paid", 1: "Free", 0: "Paid"})
                .value_counts()
                .reset_index()
                .rename(columns={"index": "Type", "free": "App Count", "count": "App Count"}))
    fp_dist.columns = ["Type", "App Count"]

    # 5. Score distribution (binned)
    score_bins = pd.cut(df["score"].dropna(), bins=[0, 1, 2, 3, 3.5, 4, 4.5, 5],
                        labels=["0-1", "1-2", "2-3", "3-3.5", "3.5-4", "4-4.5", "4.5-5"])
    score_dist = score_bins.value_counts().sort_index().reset_index()
    score_dist.columns = ["Score Range", "App Count"]

    # 6. Top 15 developers by app count
    top_devs = (df.groupby("developer")
                  .agg(app_count=("app_id", "count"),
                       avg_score=("score", "mean"),
                       avg_installs=("min_installs", "mean"))
                  .sort_values("app_count", ascending=False)
                  .head(15)
                  .reset_index()
                  .rename(columns={"developer": "Developer",
                                   "app_count": "App Count",
                                   "avg_score": "Avg Score",
                                   "avg_installs": "Avg Installs"}))

    # ── Write chart data to hidden sheet ───────────────────────────────────
    cd = _add_df_to_sheet(wb, genre_dist, "_ChartData_Genre")
    id_ = _add_df_to_sheet(wb, install_dist, "_ChartData_Installs")
    crd = _add_df_to_sheet(wb, content_dist, "_ChartData_Content")
    fpd = _add_df_to_sheet(wb, fp_dist, "_ChartData_FP")
    sd = _add_df_to_sheet(wb, score_dist, "_ChartData_Score")
    tdd = _add_df_to_sheet(wb, top_devs, "_ChartData_TopDevs")
    for sh in [cd, id_, crd, fpd, sd, tdd]:
        sh.sheet_state = "hidden"

    # ── Summary/Dashboard sheet ────────────────────────────────────────────
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    dash = wb.create_sheet("Dashboard", 0)
    dash.sheet_view.showGridLines = False
    dash.column_dimensions["A"].width = 3
    dash.column_dimensions["B"].width = 30
    dash.column_dimensions["C"].width = 20

    # Title
    dash["B2"] = "Google Play Store — 2026 Apps Dataset"
    dash["B2"].font = Font(name="Calibri", bold=True, size=18, color="1F3864")
    dash["B2"].alignment = Alignment(horizontal="left")

    stats = [
        ("Total Apps",          f"{len(df):,}"),
        ("Unique Genres",        f"{df['genre'].nunique()}"),
        ("Unique Developers",    f"{df['developer'].nunique():,}"),
        ("Avg Score",            f"{df['score'].mean():.2f} ★"),
        ("Free Apps",            f"{df['free'].sum():,} ({df['free'].mean()*100:.1f}%)"),
        ("Ad-Supported",         f"{df['ad_supported'].sum():,}"),
        ("Apps with IAP",        f"{df['in_app_purchases'].sum():,}"),
        ("Top Genre",            df['genre'].value_counts().index[0]),
        ("Highest Install App",  df.loc[df['min_installs'].idxmax(), 'title'][:40]),
    ]

    dash["B4"].font = Font(bold=True, size=12, color="FFFFFF")
    dash["B4"].fill = PatternFill("solid", fgColor="1F3864")
    dash["B4"] = "Metric"
    dash["C4"].font = Font(bold=True, size=12, color="FFFFFF")
    dash["C4"].fill = PatternFill("solid", fgColor="1F3864")
    dash["C4"] = "Value"

    for i, (metric, value) in enumerate(stats, start=5):
        dash.cell(row=i, column=2).value = metric
        dash.cell(row=i, column=2).font = Font(bold=True)
        dash.cell(row=i, column=3).value = value
        if i % 2 == 0:
            for c in [2, 3]:
                dash.cell(row=i, column=c).fill = ALT_FILL

    # Charts on dashboard
    # Genre bar chart
    _add_bar_chart(dash, cd,
                   min_row=1, max_row=len(genre_dist)+1,
                   cat_col=1, val_col=2,
                   title="Top 20 Genres by App Count",
                   anchor="E2", bar_dir="bar")

    # Install distribution bar
    _add_bar_chart(dash, id_,
                   min_row=1, max_row=len(install_dist)+1,
                   cat_col=1, val_col=2,
                   title="Install Bucket Distribution",
                   anchor="E22", bar_dir="col")

    # Content rating pie
    _add_pie_chart(dash, crd,
                   min_row=1, max_row=len(content_dist)+1,
                   cat_col=1, val_col=2,
                   title="Content Rating Mix",
                   anchor="Q2")

    # Score distribution bar
    _add_bar_chart(dash, sd,
                   min_row=1, max_row=len(score_dist)+1,
                   cat_col=1, val_col=2,
                   title="Score Distribution",
                   anchor="Q22", bar_dir="col")

    # ── Summary table sheet ─────────────────────────────────────────────────
    _add_df_to_sheet(wb, genre_dist, "Genre Summary")
    _add_df_to_sheet(wb, install_dist, "Install Summary")
    _add_df_to_sheet(wb, top_devs, "Top Developers")

    # Move Apps sheet to position 1 (after Dashboard)
    wb.move_sheet("Apps", offset=1)

    wb.save(APPS_FILE)
    print(f"  OK  apps.xlsx enhanced -- {len(df):,} rows, {len(df.columns)} columns")


# ── app_country_stats.xlsx ────────────────────────────────────────────────────

def enhance_country():
    print("Loading app_country_stats.xlsx ...")
    df = _read_data_sheet(COUNTRY_FILE, "Country Stats")

    wb = load_workbook(COUNTRY_FILE)
    if "Country Stats" in wb.sheetnames:
        ws = wb["Country Stats"]
    else:
        ws = wb.active
        ws.title = "Country Stats"

    _style_header_row(ws)
    _set_col_widths(ws, df)
    _freeze_and_filter(ws, df)
    ws.row_dimensions[1].height = 30

    # Conditional formatting on score
    if "score" in df.columns:
        sc_idx = list(df.columns).index("score") + 1
        sc_letter = get_column_letter(sc_idx)
        ws.conditional_formatting.add(
            f"{sc_letter}2:{sc_letter}{len(df)+1}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color="FF0000",
                mid_type="num",   mid_value=3,   mid_color="FFFF00",
                end_type="num",   end_value=5,   end_color="00B050",
            )
        )

    # Aggregations
    top_countries = (df.groupby("country")
                       .agg(app_count=("app_id", "count"),
                            avg_score=("score", "mean"),
                            total_installs=("min_installs", "sum"))
                       .sort_values("total_installs", ascending=False)
                       .head(20)
                       .reset_index())
    top_countries.columns = ["Country", "App Count", "Avg Score", "Total Installs"]

    score_by_country = (df.groupby("country")["score"]
                          .mean()
                          .sort_values(ascending=False)
                          .head(20)
                          .reset_index())
    score_by_country.columns = ["Country", "Avg Score"]

    # Chart data sheets (hidden)
    tc_ws = _add_df_to_sheet(wb, top_countries, "_ChartData_Countries")
    sc_ws = _add_df_to_sheet(wb, score_by_country, "_ChartData_ScoreByCountry")
    tc_ws.sheet_state = "hidden"
    sc_ws.sheet_state = "hidden"

    # Dashboard
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    dash = wb.create_sheet("Dashboard", 0)
    dash.sheet_view.showGridLines = False

    dash["B2"] = "Country Stats — 2026 Apps"
    dash["B2"].font = Font(name="Calibri", bold=True, size=18, color="1F3864")

    stats = [
        ("Total Records",     f"{len(df):,}"),
        ("Unique Countries",  f"{df['country'].nunique()}"),
        ("Unique Apps",       f"{df['app_id'].nunique():,}"),
        ("Top Country",       df['country'].value_counts().index[0]),
    ]
    dash["B4"] = "Metric"; dash["B4"].fill = PatternFill("solid", fgColor="1F3864")
    dash["B4"].font = Font(bold=True, color="FFFFFF")
    dash["C4"] = "Value";  dash["C4"].fill = PatternFill("solid", fgColor="1F3864")
    dash["C4"].font = Font(bold=True, color="FFFFFF")
    for i, (m, v) in enumerate(stats, 5):
        dash.cell(row=i, column=2).value = m
        dash.cell(row=i, column=2).font = Font(bold=True)
        dash.cell(row=i, column=3).value = v

    _add_bar_chart(dash, tc_ws,
                   min_row=1, max_row=len(top_countries)+1,
                   cat_col=1, val_col=4,
                   title="Top 20 Countries — Total Installs",
                   anchor="E2", bar_dir="bar")

    _add_bar_chart(dash, sc_ws,
                   min_row=1, max_row=len(score_by_country)+1,
                   cat_col=1, val_col=2,
                   title="Top 20 Countries — Avg Score",
                   anchor="E22", bar_dir="bar")

    _add_df_to_sheet(wb, top_countries, "Country Summary")
    wb.move_sheet("Country Stats", offset=1)

    wb.save(COUNTRY_FILE)
    print(f"  OK  app_country_stats.xlsx enhanced -- {len(df):,} rows")


# ── discovery_signals.xlsx ────────────────────────────────────────────────────

def enhance_discovery():
    print("Loading discovery_signals.xlsx ...")
    df = _read_data_sheet(DISCOVERY_FILE, "Discovery Signals")

    wb = load_workbook(DISCOVERY_FILE)
    if "Discovery Signals" in wb.sheetnames:
        ws = wb["Discovery Signals"]
    else:
        ws = wb.active
        ws.title = "Discovery Signals"

    _style_header_row(ws)
    _set_col_widths(ws, df)
    _freeze_and_filter(ws, df)
    ws.row_dimensions[1].height = 30

    # Aggregations
    top_keywords = (df.groupby("keyword")
                      .size()
                      .sort_values(ascending=False)
                      .head(30)
                      .reset_index()
                      .rename(columns={0: "frequency"}))
    top_keywords.columns = ["Keyword", "Frequency"]

    source_dist = (df.groupby("source").size().reset_index()
                     .rename(columns={0: "count"}))
    source_dist.columns = ["Source", "Count"]

    top_categories = (df.groupby("category").size()
                        .sort_values(ascending=False).head(20)
                        .reset_index().rename(columns={0: "count"}))
    top_categories.columns = ["Category", "Count"]

    country_dist = (df.groupby("country").size()
                      .sort_values(ascending=False).head(20)
                      .reset_index().rename(columns={0: "count"}))
    country_dist.columns = ["Country", "Signals"]

    # Chart data sheets (hidden)
    kw_ws  = _add_df_to_sheet(wb, top_keywords,    "_ChartData_Keywords")
    src_ws = _add_df_to_sheet(wb, source_dist,     "_ChartData_Source")
    cat_ws = _add_df_to_sheet(wb, top_categories,  "_ChartData_DiscCat")
    co_ws  = _add_df_to_sheet(wb, country_dist,    "_ChartData_DiscCountry")
    for sh in [kw_ws, src_ws, cat_ws, co_ws]:
        sh.sheet_state = "hidden"

    # Dashboard
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    dash = wb.create_sheet("Dashboard", 0)
    dash.sheet_view.showGridLines = False

    dash["B2"] = "Discovery Signals — 2026 Apps"
    dash["B2"].font = Font(name="Calibri", bold=True, size=18, color="1F3864")

    stats = [
        ("Total Signals",       f"{len(df):,}"),
        ("Unique Apps",         f"{df['app_id'].nunique():,}"),
        ("Unique Keywords",     f"{df['keyword'].dropna().nunique():,}"),
        ("Unique Countries",    f"{df['country'].dropna().nunique()}"),
        ("Unique Sources",      f"{df['source'].dropna().nunique()}"),
        ("Top Keyword",         df['keyword'].value_counts().index[0] if df['keyword'].notna().any() else "N/A"),
    ]
    dash["B4"] = "Metric"; dash["B4"].fill = PatternFill("solid", fgColor="1F3864")
    dash["B4"].font = Font(bold=True, color="FFFFFF")
    dash["C4"] = "Value";  dash["C4"].fill = PatternFill("solid", fgColor="1F3864")
    dash["C4"].font = Font(bold=True, color="FFFFFF")
    for i, (m, v) in enumerate(stats, 5):
        dash.cell(row=i, column=2).value = m
        dash.cell(row=i, column=2).font = Font(bold=True)
        dash.cell(row=i, column=3).value = v

    _add_bar_chart(dash, kw_ws,
                   min_row=1, max_row=len(top_keywords)+1,
                   cat_col=1, val_col=2,
                   title="Top 30 Keywords by Frequency",
                   anchor="E2", bar_dir="bar")

    _add_bar_chart(dash, cat_ws,
                   min_row=1, max_row=len(top_categories)+1,
                   cat_col=1, val_col=2,
                   title="Top 20 Discovery Categories",
                   anchor="E22", bar_dir="bar")

    _add_pie_chart(dash, src_ws,
                   min_row=1, max_row=len(source_dist)+1,
                   cat_col=1, val_col=2,
                   title="Signal Source Mix",
                   anchor="Q2")

    _add_bar_chart(dash, co_ws,
                   min_row=1, max_row=len(country_dist)+1,
                   cat_col=1, val_col=2,
                   title="Top 20 Countries — Signal Count",
                   anchor="Q22", bar_dir="bar")

    _add_df_to_sheet(wb, top_keywords,   "Keyword Summary")
    _add_df_to_sheet(wb, source_dist,    "Source Summary")
    wb.move_sheet("Discovery Signals", offset=1)

    wb.save(DISCOVERY_FILE)
    print(f"  OK  discovery_signals.xlsx enhanced -- {len(df):,} rows")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    enhance_apps()
    enhance_country()
    enhance_discovery()
    print("\nAll three Excel files enhanced successfully.")
